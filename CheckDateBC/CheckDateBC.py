#!/usr/bin/env python3
"""
Excel Date Validator & Auto-Fixer (openpyxl-only)
- Normalises target columns to dd/mm/YYYY.
- Auto-fixes safe cases (e.g., 29.09.2025 -> 29/09/2025; Excel serials).
- Reports unfixable values with precise coordinates.
- NEW: --last-row to limit processing window (e.g., --last-row 73).

Usage:
  python check_dates_openpyxl.py --in "input.xlsx" --last-row 73
"""

import argparse
from datetime import datetime
from typing import Optional, Tuple, List, Union
from openpyxl import load_workbook, Workbook
from openpyxl.utils.datetime import from_excel, CALENDAR_WINDOWS_1900

TARGET_DATE_FORMAT = "%d/%m/%Y"
DEFAULT_YEAR_MIN = 2000
DEFAULT_YEAR_MAX = 2099

TARGET_COLUMNS = {
    "DATE",
    "DATE OF 2ND VISIT TO CLEAN REMAINING TAGS",
    "DATE OF 3RD VISIT TO CLEAN REMAINING TAGS",
    "DATE OF COMPLETING ALL SHADOWS (TAGS) ON THIS UNIT EXCLUDING ACID ETCHED SHADOWS",
}

def is_target_header(text: Union[str, None]) -> bool:
    if text is None:
        return False
    return text.strip().lower() in {c.lower() for c in TARGET_COLUMNS}

def try_excel_serial(value: Union[int, float]) -> Optional[datetime]:
    try:
        return from_excel(value, CALENDAR_WINDOWS_1900)
    except Exception:
        return None

def normalise_date(raw, year_min: int, year_max: int) -> Tuple[Optional[str], str]:
    """
    Return (fixed_value_or_None, status) where status ∈ {"ok","fixed","empty","invalid"}.
    """
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        return (None, "empty")

    if isinstance(raw, datetime):
        y = raw.year
        if year_min <= y <= year_max:
            return (raw.strftime(TARGET_DATE_FORMAT), "fixed")
        return (None, "invalid")

    if isinstance(raw, (int, float)):
        dt = try_excel_serial(raw)
        if dt:
            if year_min <= dt.year <= year_max:
                return (dt.strftime(TARGET_DATE_FORMAT), "fixed")
            return (None, "invalid")

    s = str(raw).strip()
    if s == "":
        return (None, "empty")

    s2 = (
        s.replace(".", "/")
         .replace("-", "/")
         .replace("\\", "/")
         .replace(" ", "/")
    )
    while "//" in s2:
        s2 = s2.replace("//", "/")

    parts = s2.split("/")
    if len(parts) != 3:
        return (None, "invalid")

    d, m, y = parts[0].strip(), parts[1].strip(), parts[2].strip()
    if not (d.isdigit() and m.isdigit() and y.isdigit()):
        return (None, "invalid")
    if len(y) != 4:
        return (None, "invalid")

    d_i, m_i, y_i = int(d), int(m), int(y)
    if not (year_min <= y_i <= year_max):
        return (None, "invalid")

    try:
        dt = datetime(y_i, m_i, d_i)
    except ValueError:
        return (None, "invalid")

    fixed = dt.strftime(TARGET_DATE_FORMAT)
    original_normalised = f"{str(d_i).zfill(2)}/{str(m_i).zfill(2)}/{y}"
    return (fixed, "ok" if fixed == original_normalised else "fixed")

def process_workbook(path_in: str, path_out: str, report_out: str,
                     year_min: int, year_max: int, last_row: Optional[int]) -> Tuple[int, int]:
    wb = load_workbook(path_in, data_only=True)
    issues: List[dict] = []
    fixed_count = 0

    for sheet_idx, ws in enumerate(wb.worksheets, start=1):
        # Headers assumed on row 1
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            hdr = cell.value if isinstance(cell.value, str) else (str(cell.value) if cell.value is not None else None)
            headers[col_idx] = hdr

        target_col_indexes = [c for c, h in headers.items() if is_target_header(h)]
        if not target_col_indexes:
            continue

        # Determine processing window
        data_last_row = min(ws.max_row, last_row) if last_row else ws.max_row

        for row_idx in range(2, data_last_row + 1):
            for col_idx in target_col_indexes:
                cell = ws.cell(row=row_idx, column=col_idx)
                raw = cell.value

                fixed, status = normalise_date(raw, year_min, year_max)
                if status in ("ok", "fixed"):
                    before = "" if raw is None else str(raw).strip()
                    cell.value = fixed
                    if status == "fixed" and fixed != before:
                        fixed_count += 1
                elif status == "empty":
                    continue
                else:
                    issues.append({
                        "Sheet Index (1-based)": sheet_idx,
                        "Sheet Name": ws.title,
                        "Column Index (1-based)": col_idx,
                        "Column Name": headers.get(col_idx, f"Column {col_idx}"),
                        "Row (Excel numbering)": row_idx,
                        "Original Value": raw,
                        "Issue": "Unparseable or out-of-range date",
                    })

    wb.save(path_out)

    rep_wb = Workbook()
    rep_ws = rep_wb.active
    rep_ws.title = "Issues"
    if issues:
        cols = list(issues[0].keys())
        rep_ws.append(cols)
        for rec in issues:
            rep_ws.append([rec.get(k) for k in cols])
    else:
        rep_ws.append(["Message"])
        rep_ws.append(["No issues detected in target columns."])
    rep_wb.save(report_out)

    return fixed_count, len(issues)

def main():
    ap = argparse.ArgumentParser(description="Validate and fix Excel date columns (openpyxl-only).")
    ap.add_argument("--in", dest="inp", required=True, help="Input .xlsx file")
    ap.add_argument("--out", dest="out", default=None, help="Output corrected .xlsx (default: append _corrected)")
    ap.add_argument("--report", dest="report", default=None, help="Issues report .xlsx (default: append _report)")
    ap.add_argument("--year-min", type=int, default=DEFAULT_YEAR_MIN, help="Minimum valid year (default 2000)")
    ap.add_argument("--year-max", type=int, default=DEFAULT_YEAR_MAX, help="Maximum valid year (default 2099)")
    ap.add_argument("--last-row", type=int, default=None, help="Only process rows up to this Excel row number (inclusive)")
    args = ap.parse_args()

    inp = args.inp
    out = args.out or (inp[:-5] + "_corrected.xlsx" if inp.lower().endswith(".xlsx") else inp + "_corrected.xlsx")
    report = args.report or (inp[:-5] + "_report.xlsx" if inp.lower().endswith(".xlsx") else inp + "_report.xlsx")

    fixed, invalid = process_workbook(inp, out, report, args.year_min, args.year_max, args.last_row)
    print(f"Done. Fixed: {fixed}   Invalid (flagged in report): {invalid}")
    print(f"Corrected workbook: {out}")
    print(f"Issues report: {report}")

if __name__ == "__main__":
    main()