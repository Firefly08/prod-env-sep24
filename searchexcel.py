import os
import shutil
import datetime as dt
from typing import Tuple

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl is required for editing .xlsx files. pip install openpyxl")

# --- CONFIG ---
folder_path = r"C:\Users\laho2188\Downloads\Enter"
search_text_default = "18/07/225"  # used if you just press Enter at the prompt

# --- OUTPUT SETUP ---
ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = os.path.join(folder_path, f"search_replace_report_{ts}.txt")

def log(f, msg):
    print(msg, end="" if msg.endswith("\n") else "\n")
    f.write(msg if msg.endswith("\n") else msg + "\n")

def search_file_xlslike(file_path: str) -> Tuple[int, list]:
    """
    Search only (no edit) for legacy .xls or any file we can't open with openpyxl.
    Returns (match_count, match_details[])
    """
    matches = []
    try:
        xls = pd.ExcelFile(file_path)  # uses xlrd/odf/pyxlsb depending on install
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
            for r_idx, row in df.iterrows():
                for c_idx, val in row.items():
                    if val is not None and str(val).strip() == search_text:
                        matches.append((sheet, r_idx + 1, c_idx + 1))
    except Exception as e:
        matches.append(("__ERROR__", -1, -1))
        raise e
    return len(matches), matches

def search_and_replace_xlsx(file_path: str, old: str, new: str) -> Tuple[int, list, int]:
    """
    Search and replace within .xlsx using openpyxl.
    Returns (match_count, match_details[], replace_count)
    """
    wb = load_workbook(file_path)
    matches = []
    replaces = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v is not None and str(v).strip() == old:
                    matches.append((ws.title, cell.row, cell.column))
                    # Preserve type if cell was string; otherwise write as string (explicit requirement is text match)
                    cell.value = new
                    replaces += 1
    return len(matches), matches, replaces, wb

# --- INTERACTION ---
print(f"Searching for text in: {folder_path}")
search_text = input(f"Text to find (default '{search_text_default}'): ").strip() or search_text_default
replacement_text = input("How do you want to fix it? Enter replacement text: ").strip()

if replacement_text == "":
    raise SystemExit("No replacement text provided. Aborting to be safe.")

with open(output_file, "w", encoding="utf-8") as f:
    log(f, f"=== Excel Search & Replace Report ===")
    log(f, f"Run at: {dt.datetime.now().isoformat(timespec='seconds')}")
    log(f, f"Folder: {folder_path}")
    log(f, f"Find: '{search_text}'  ->  Replace with: '{replacement_text}'")
    log(f, "")

    total_files = 0
    total_matches = 0
    total_replacements = 0
    edited_files = 0
    skipped_edits_xls = 0
    errors = 0

    for filename in os.listdir(folder_path):
        if not (filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls")):
            continue

        total_files += 1
        file_path = os.path.join(folder_path, filename)

        try:
            if filename.lower().endswith(".xlsx"):
                # Search & replace using openpyxl
                m, details, r, wb = search_and_replace_xlsx(file_path, search_text, replacement_text)
                total_matches += m

                if m > 0:
                    # Backup before writing changes
                    backup_path = os.path.join(folder_path, f"BACKUP_{filename}_{ts}.xlsx")
                    shutil.copy2(file_path, backup_path)

                    wb.save(file_path)
                    edited_files += 1
                    total_replacements += r

                    log(f, f"✅ {filename}: {m} match(es), {r} replacement(s) committed.")
                    for (sheet, row, col) in details[:25]:
                        log(f, f"   - {sheet}: Row {row}, Col {col}")
                    if len(details) > 25:
                        log(f, f"   ...and {len(details)-25} more cells.")
                else:
                    log(f, f"— {filename}: no matches.")

            else:
                # .xls — search only, no edits
                try:
                    m, details = search_file_xlslike(file_path)
                    total_matches += m
                    if m > 0:
                        skipped_edits_xls += 1
                        log(f, f"🟨 {filename}: {m} match(es) FOUND, but not edited (legacy .xls).")
                        for (sheet, row, col) in details[:25]:
                            if sheet == "__ERROR__":
                                break
                            log(f, f"   - {sheet}: Row {row}, Col {col}")
                        if len(details) > 25:
                            log(f, f"   ...and {len(details)-25} more cells.")
                    else:
                        log(f, f"— {filename}: no matches.")
                except Exception as se:
                    errors += 1
                    log(f, f"⚠️  {filename}: search failed. Reason: {se}")

        except Exception as e:
            errors += 1
            log(f, f"⚠️  {filename}: processing failed. Reason: {e}")

    log(f, "")
    log(f, "=== Summary ===")
    log(f, f"Files scanned: {total_files}")
    log(f, f"Total matches: {total_matches}")
    log(f, f"Total replacements committed: {total_replacements}")
    log(f, f"Files edited (.xlsx): {edited_files}")
    if skipped_edits_xls:
        log(f, f"Legacy .xls with matches (not edited): {skipped_edits_xls}  "
               "→ Convert to .xlsx to allow edits.")
    if errors:
        log(f, f"Files with errors: {errors}")
    log(f, f"Report saved: {output_file}")

print(f"\nDone. Full report: {output_file}")
