#!/usr/bin/env python3
# v3.5 — fixes EmptyCell by using values_only + index-based coords;
#        scans only three date columns; supports serial dates (epoch-aware)
# v3.4 — robust date scan (serial dates + header-restricted)
# v3.3 — DateMatches scan
# v3.2 — optional leading "Inspected" sheet

from __future__ import annotations
import argparse, glob as _glob, os
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional, Set
from datetime import datetime, date
import warnings
warnings.filterwarnings("ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning, module="openpyxl")
import pandas as pd

try:
    import openpyxl  # type: ignore
    from openpyxl.utils.datetime import from_excel  # type: ignore
    from openpyxl.utils import get_column_letter    # type: ignore
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

EXPECTED_EFFECTIVE_SHEET_COUNT = 13
EXPECTED_REPEAT_SHEETS = [f"Repeat Attacks {i}" for i in range(1, 13)]

# --- Date scan targets (kept as UK dd/mm/yyyy strings) ---
TARGET_DATE_STRS = ["10/12/2025", "04/11/2026"]

# --- Column headers to scan only ---
TARGET_DATE_HEADERS: Set[str] = {
    "Date",
    "Date of 2nd visit to clean remaining tags",
    "Date of 3rd visit to clean remaining tags",
}

@dataclass
class TableMeta:
    name: str
    sheet: str
    ref: str

@dataclass
class FileSchema:
    path: str
    sheet_names: List[str] = field(default_factory=list)
    tables: Dict[str, str] = field(default_factory=dict)            # table name -> sheet
    tables_meta: Dict[str, TableMeta] = field(default_factory=dict)  # table name -> meta

def discover_schema(path: str) -> FileSchema:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for this validator but is not available.")
    schema = FileSchema(path=path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        schema.sheet_names = [ws.title for ws in wb.worksheets]
        for ws in wb.worksheets:
            # Handle both dict-like and list-like table containers
            try:
                items_iter = ws.tables.items()
            except Exception:
                try:
                    items_iter = [(getattr(t, "name", str(i)), t) for i, t in enumerate(ws.tables)]
                except Exception:
                    items_iter = []
            for tname, tval in items_iter:
                tname = str(tname)
                ref = getattr(tval, "ref", None)
                if ref is None and isinstance(tval, str):
                    ref = tval
                if ref is None:
                    ref = ""
                schema.tables[tname] = ws.title
                schema.tables_meta[tname] = TableMeta(name=tname, sheet=ws.title, ref=ref)
    finally:
        try: wb.close()
        except Exception: pass
    return schema

def validate_file(schema: FileSchema) -> Tuple[pd.DataFrame, Dict[str, bool]]:
    issues = []

    sheet_names = schema.sheet_names
    has_leading_inspected = bool(sheet_names) and sheet_names[0].strip() == "Inspected"
    # Accept 13 sheets OR 14 with leading "Inspected"
    total_ok = (len(sheet_names) == 13) or (has_leading_inspected and len(sheet_names) == 14)
    if not total_ok:
        issues.append({
            "file": os.path.basename(schema.path),
            "sheet": "-",
            "issue": "Sheet count mismatch",
            "detail": ("found {} (allowed: 13) or 14 with first sheet named 'Inspected'"
                       .format(len(sheet_names)))
        })

    # Build effective list (skip 'Inspected' if first)
    effective = sheet_names[1:] if has_leading_inspected else sheet_names

    # Must have exactly 13 effective sheets
    if len(effective) != EXPECTED_EFFECTIVE_SHEET_COUNT:
        issues.append({
            "file": os.path.basename(schema.path),
            "sheet": "-",
            "issue": "Effective sheet count mismatch",
            "detail": f"effective={len(effective)}; expected {EXPECTED_EFFECTIVE_SHEET_COUNT}"
        })

    # Check names for effective sheets 2..13
    for idx, expected_name in enumerate(EXPECTED_REPEAT_SHEETS, start=2):
        actual = effective[idx - 1] if len(effective) >= idx else None
        if actual != expected_name:
            issues.append({
                "file": os.path.basename(schema.path),
                "sheet": actual if actual is not None else "-",
                "issue": "Sheet name mismatch",
                "detail": f"effective position {idx}: expected '{expected_name}', found '{actual}'"
            })

    # Table placement:
    # effective[0] -> TagsRA0; effective[1..12] -> TagsRA1..12
    def exp_sheet_for_index(eff_idx: int) -> Optional[str]:
        if eff_idx == 1:
            return effective[0] if effective else None
        else:
            i = eff_idx - 1  # 1..12
            return f"Repeat Attacks {i}"

    # Validate table presence and sheet assignment
    for eff_idx in range(1, 14):  # 1..13
        expected_table = "TagsRA0" if eff_idx == 1 else f"TagsRA{eff_idx-1}"
        expected_sheet_name = exp_sheet_for_index(eff_idx)
        actual_table_sheet = schema.tables.get(expected_table)
        if actual_table_sheet is None:
            issues.append({
                "file": os.path.basename(schema.path),
                "sheet": expected_sheet_name if expected_sheet_name else "-",
                "issue": "Missing table",
                "detail": f"expected table '{expected_table}' not found"
            })
        else:
            if expected_sheet_name and actual_table_sheet != expected_sheet_name:
                issues.append({
                    "file": os.path.basename(schema.path),
                    "sheet": actual_table_sheet,
                    "issue": "Table on wrong sheet",
                    "detail": f"table '{expected_table}' expected on sheet '{expected_sheet_name}', found on '{actual_table_sheet}'"
                })

    detail_df = pd.DataFrame(issues, columns=["file", "sheet", "issue", "detail"])
    status = {
        "sheet_count_ok": total_ok,
        "effective_count_ok": len(detail_df[detail_df["issue"] == "Effective sheet count mismatch"]) == 0,
        "repeat_sheets_ok": len(detail_df[detail_df["issue"] == "Sheet name mismatch"]) == 0,
        "tables_present_ok": len(detail_df[detail_df["issue"] == "Missing table"]) == 0,
        "tables_on_expected_sheet_ok": len(detail_df[detail_df["issue"] == "Table on wrong sheet"]) == 0,
    }
    return detail_df, status

# ---- Date scan helpers -------------------------------------------------------

def _parse_targets() -> Tuple[set, set]:
    """Build two sets of datetime.date for robust comparison (dd/mm and mm/dd)."""
    dmy, mdy = set(), set()
    for s in TARGET_DATE_STRS:
        s = s.strip()
        try:
            dmy.add(datetime.strptime(s, "%d/%m/%Y").date())
        except Exception:
            pass
        try:
            mdy.add(datetime.strptime(s, "%m/%d/%Y").date())
        except Exception:
            pass
    return dmy, mdy

_DMY_TARGETS, _MDY_TARGETS = _parse_targets()
_TEXT_TARGETS = set(TARGET_DATE_STRS)

def _coerce_to_date(val, wb_epoch) -> Optional[date]:
    """Coerce to date from datetime/date/str or Excel serial (int/float)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            return from_excel(val, wb_epoch).date()
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
    return None

def _detect_target_columns(ws) -> Set[int]:
    """
    Identify columns whose header matches any TARGET_DATE_HEADERS.
    Scans the top 10 rows using values_only to avoid EmptyCell objects.
    Returns 1-based column indexes.
    """
    targets_lower = {h.lower() for h in TARGET_DATE_HEADERS}
    target_cols: Set[int] = set()
    max_scan_rows = min(ws.max_row or 0, 10)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        if row is None:
            continue
        for col_idx, v in enumerate(row, start=1):
            if isinstance(v, str) and v.strip().lower() in targets_lower:
                target_cols.add(col_idx)
    return target_cols

def find_date_hits(xlsx_path: str) -> pd.DataFrame:
    """
    Scan only target date columns (by header match) for 10/12/2025 and 04/11/2026.
    Uses values_only=True to avoid EmptyCell; computes coordinates from indexes.
    Returns DataFrame[file, sheet, cell, line, value, header].
    """
    if not _HAS_OPENPYXL:
        return pd.DataFrame([{
            "file": os.path.basename(xlsx_path),
            "sheet": "-",
            "cell": "-",
            "line": "-",
            "value": "openpyxl not available",
            "header": "-"
        }], columns=["file", "sheet", "cell", "line", "value", "header"])

    rows = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        wb_epoch = wb.epoch
        for ws in wb.worksheets:
            target_cols = _detect_target_columns(ws)
            if not target_cols:
                continue

            # Map col -> header text (first occurrence in top 10 rows)
            header_map: Dict[int, str] = {}
            max_scan_rows = min(ws.max_row or 0, 10)
            for row in ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True):
                if row is None:
                    continue
                for col_idx, v in enumerate(row, start=1):
                    if col_idx in target_cols and col_idx not in header_map and isinstance(v, str):
                        header_map[col_idx] = v.strip()

            # Data scan
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row is None:
                    continue
                for col_idx in target_cols:
                    val = row[col_idx - 1] if (col_idx - 1) < len(row) else None

                    d = _coerce_to_date(val, wb_epoch)
                    hit = (d is not None and (d in _DMY_TARGETS or d in _MDY_TARGETS)) \
                          or (isinstance(val, str) and val.strip() in _TEXT_TARGETS)

                    if hit:
                        coord = f"{get_column_letter(col_idx)}{row_idx}"
                        rows.append({
                            "file": os.path.basename(xlsx_path),
                            "sheet": ws.title,
                            "cell": coord,
                            "line": row_idx,
                            "value": (d.strftime("%d/%m/%Y") if d is not None
                                      else (val.strip() if isinstance(val, str) else str(val))),
                            "header": header_map.get(col_idx, "-"),
                        })
    finally:
        try: wb.close()
        except Exception: pass

    return pd.DataFrame(rows, columns=["file", "sheet", "cell", "line", "value", "header"])

# -----------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Validate Excel files and scan target date columns for specific dates."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--files", nargs="+", help="List of Excel files to validate.")
    group.add_argument("--glob", dest="glob_pattern", help="Glob pattern, e.g., 'data/*.xlsx'.")
    parser.add_argument("--out", default=os.path.join("results", "schema_report.xlsx"),
                        help="Output report path (Excel). Default: results/schema_report.xlsx")
    args = parser.parse_args()

    files = args.files if args.files else sorted(_glob.glob(args.glob_pattern))
    if not files:
        raise SystemExit("No input files. Provide --files or --glob.")

    out_dir = os.path.dirname(os.path.abspath(args.out))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    all_details: List[pd.DataFrame] = []
    all_date_hits: List[pd.DataFrame] = []
    summary_rows = []

    for f in files:
        if not os.path.exists(f):
            raise SystemExit(f"File not found: {f}")
        if not _HAS_OPENPYXL:
            raise SystemExit("openpyxl not available; cannot validate tables/sheets.")

        try:
            schema = discover_schema(f)
            detail_df, status = validate_file(schema)
        except Exception as e:
            detail_df = pd.DataFrame([{
                "file": os.path.basename(f),
                "sheet": "-",
                "issue": "Read error",
                "detail": str(e)
            }], columns=["file", "sheet", "issue", "detail"])
            status = {
                "sheet_count_ok": False,
                "effective_count_ok": False,
                "repeat_sheets_ok": False,
                "tables_present_ok": False,
                "tables_on_expected_sheet_ok": False,
            }

        # Date scan
        try:
            date_hits = find_date_hits(f)
        except Exception as e:
            date_hits = pd.DataFrame([{
                "file": os.path.basename(f),
                "sheet": "-",
                "cell": "-",
                "line": "-",
                "value": f"Scan error: {e}",
                "header": "-"
            }], columns=["file", "sheet", "cell", "line", "value", "header"])

        all_details.append(detail_df if not detail_df.empty else pd.DataFrame(
            [{"file": os.path.basename(f), "sheet": "-", "issue": "None", "detail": "No issues found"}],
            columns=["file", "sheet", "issue", "detail"]
        ))
        all_date_hits.append(date_hits)

        overall_pass = all(status.values())
        summary_rows.append({
            "file": os.path.basename(f),
            **status,
            "overall": "PASS" if overall_pass else "FAIL",
            "issues": int(detail_df.shape[0])
        })

    summary_df = pd.DataFrame(summary_rows, columns=[
        "file", "sheet_count_ok", "effective_count_ok", "repeat_sheets_ok",
        "tables_present_ok", "tables_on_expected_sheet_ok", "overall", "issues"
    ])
    details_combined = pd.concat(all_details, ignore_index=True) if all_details else pd.DataFrame(
        columns=["file", "sheet", "issue", "detail"]
    )
    date_hits_combined = pd.concat(all_date_hits, ignore_index=True) if all_date_hits else pd.DataFrame(
        columns=["file","sheet","cell","line","value","header"]
    )
    if date_hits_combined.empty:
        date_hits_combined = pd.DataFrame(
            [{"file": "-", "sheet": "-", "cell": "-", "line": "-", "value": "No matches found", "header": "-"}],
            columns=["file","sheet","cell","line","value","header"]
        )

    with pd.ExcelWriter(args.out, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        details_combined.to_excel(writer, index=False, sheet_name="Details")
        date_hits_combined.to_excel(writer, index=False, sheet_name="DateMatches")

    print(f"Report written to: {os.path.abspath(args.out)}")
    print("Validation complete: optional leading 'Inspected'; 13 effective sheets; table placement TagsRA0..TagsRA12.")
    print(f"Date scan targets: {', '.join(TARGET_DATE_STRS)}")
    print(f"Matches: {0 if (len(date_hits_combined)==1 and date_hits_combined.iloc[0]['file']=='-') else len(date_hits_combined)} (see 'DateMatches')")

if __name__ == "__main__":
    raise SystemExit(main())
