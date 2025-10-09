import os
import sys
import datetime as dt
from typing import List

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl is required to edit .xlsx files. Install with: pip install openpyxl")

# ========== CONFIG ==========
folder_path = r"C:\Users\laho2188\Downloads\Enter"
date_to_find_default = "18/07/225"  # default if you just press Enter

# ========== UTIL ==========
def nowstamp():
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def ts_slug():
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")

def log(wfile, msg):
    print(msg)
    wfile.write(msg + ("\n" if not msg.endswith("\n") else ""))

def safe_str(x):
    return "" if x is None else str(x).strip()

# ========== SEARCH HELPERS ==========
def search_xls_or_xlsx_readonly(file_path: str, target: str) -> List[tuple]:
    """Read-only exact-match search across all sheets/cells. Returns list of (sheet, row, col)."""
    matches = []
    xls = pd.ExcelFile(file_path)  # let pandas pick the engine
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
        for r_idx, row in df.iterrows():
            for c_idx, val in row.items():
                if safe_str(val) == target:
                    matches.append((sheet, r_idx + 1, c_idx + 1))
    return matches

def search_and_replace_xlsx_inplace(file_path: str, old: str, new: str) -> List[tuple]:
    """Edit .xlsx in-place using openpyxl. Returns list of (sheet, row, col) replaced."""
    wb = load_workbook(file_path)
    replaced = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None and safe_str(cell.value) == old:
                    replaced.append((ws.title, cell.row, cell.column))
                    cell.value = new
    if replaced:
        wb.save(file_path)
    return replaced

# ========== MAIN ==========
if __name__ == "__main__":
    print(f"Scanning folder: {folder_path}")
    date_to_find = input(f"Text to find (default '{date_to_find_default}'): ").strip() or date_to_find_default

    report_path = os.path.join(folder_path, f"excel_search_report_{ts_slug()}.txt")
    total_files = 0
    total_matches = 0
    file_hits = 0
    errors = 0
    per_file_results = {}  # filename -> list of (sheet,row,col)
    unreadable = []        # files that couldn't be read

    # Phase 1 — Search & Report
    with open(report_path, "w", encoding="utf-8") as f:
        log(f, f"=== Excel Search Report ===")
        log(f, f"Run: {nowstamp()}")
        log(f, f"Folder: {folder_path}")
        log(f, f"Exact match search for: '{date_to_find}'")
        log(f, "")

        for filename in os.listdir(folder_path):
            if not (filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls")):
                continue

            total_files += 1
            file_path = os.path.join(folder_path, filename)

            try:
                matches = search_xls_or_xlsx_readonly(file_path, date_to_find)
                if matches:
                    file_hits += 1
                    total_matches += len(matches)
                    per_file_results[filename] = matches
                    log(f, f"✅ {filename}: {len(matches)} match(es)")
                    for (sheet, r, c) in matches[:25]:
                        log(f, f"   - {sheet}: Row {r}, Column {c}")
                    if len(matches) > 25:
                        log(f, f"   ...and {len(matches)-25} more cells.")
                else:
                    log(f, f"— {filename}: no matches.")
            except Exception as e:
                errors += 1
                unreadable.append((filename, str(e)))
                log(f, f"⚠️  {filename}: could not be read. Reason: {e}")

        # Summary
        log(f, "")
        log(f, "=== Summary ===")
        log(f, f"Files scanned: {total_files}")
        log(f, f"Files with matches: {file_hits}")
        log(f, f"Total matches: {total_matches}")
        if unreadable:
            log(f, f"Unreadable files: {len(unreadable)}")
            for fn, reason in unreadable[:10]:
                log(f, f"   - {fn}: {reason}")
            if len(unreadable) > 10:
                log(f, f"   ...and {len(unreadable)-10} more.")
        log(f, f"Report saved: {report_path}")

    print("\n--- Phase 1 complete. ---")
    print(f"TXT report path: {report_path}")

    # If nothing found, exit cleanly
    if total_matches == 0:
        print("No matches found. No changes required. Exiting.")
        sys.exit(0)

    # Phase 2 — Optional Fix (only offered if matches exist)
    proceed = input("\nMatches found. Do you want to fix/replace them in .xlsx files now? (y/N): ").strip().lower()
    if proceed not in ("y", "yes"):
        print("Acknowledged. No changes made. Exiting.")
        sys.exit(0)

    replacement = input("How do you want to fix it? Enter replacement text: ").strip()
    if replacement == "":
        print("No replacement provided. Exiting without changes.")
        sys.exit(0)

    # Replace in-place (no backups)
    with open(report_path, "a", encoding="utf-8") as f:
        log(f, "")
        log(f, "=== Replace Operation (No Backups) ===")
        log(f, f"Replacing '{date_to_find}' -> '{replacement}'")
        log(f, f"Started: {nowstamp()}")
        log(f, "")

        edited_files = 0
        total_replacements = 0
        skipped_non_xlsx = 0
        replace_errors = 0

        for filename, matches in per_file_results.items():
            file_path = os.path.join(folder_path, filename)
            if not filename.lower().endswith(".xlsx"):
                skipped_non_xlsx += 1
                log(f, f"🟨 {filename}: has matches but is not .xlsx. Skipped (convert to .xlsx to edit).")
                continue

            try:
                replaced_cells = search_and_replace_xlsx_inplace(file_path, date_to_find, replacement)
                if replaced_cells:
                    edited_files += 1
                    total_replacements += len(replaced_cells)
                    log(f, f"✅ {filename}: {len(replaced_cells)} replacement(s) committed.")
                else:
                    log(f, f"— {filename}: nothing to replace at commit time (unexpected).")
            except Exception as e:
                replace_errors += 1
                log(f, f"⚠️  {filename}: replace failed. Reason: {e}")

        log(f, "")
        log(f, "=== Replace Summary ===")
        log(f, f"Files edited (.xlsx): {edited_files}")
        log(f, f"Total replacements committed: {total_replacements}")
        if skipped_non_xlsx:
            log(f, f"Files with matches skipped (not .xlsx): {skipped_non_xlsx}")
        if replace_errors:
            log(f, f"Files with replace errors: {replace_errors}")
        log(f, f"Finished: {nowstamp()}")
        log(f, f"Updated report: {report_path}")

    print("\n--- Phase 2 complete. ---")
    print(f"Updated report: {report_path}")
